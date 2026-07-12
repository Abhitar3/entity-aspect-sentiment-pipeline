import argparse
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def excel_to_posts_json(input_xlsx: Path, output_json: Path, column_name: str, sheet_name: str | None = None) -> int:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is not installed. Install it with: pip install openpyxl"
        )

    if not input_xlsx.exists():
        raise FileNotFoundError(f"Excel file not found: {input_xlsx}")

    workbook = load_workbook(filename=input_xlsx, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(worksheet.iter_rows(max_row=1))]
    if column_name not in headers:
        raise ValueError(
            f"Column '{column_name}' not found in sheet '{worksheet.title}'. "
            f"Available columns: {headers}"
        )

    col_index = headers.index(column_name)

    posts = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        value = row[col_index] if col_index < len(row) else None
        if value is None:
            continue
        text = str(value).strip()
        if text:
            posts.append(text)

    data = {"posts": posts}
    output_json.parent.mkdir(parents=True, exist_ok=True)
    with output_json.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    return len(posts)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert an Excel sheet column to pipeline input JSON.")
    parser.add_argument("input", help="Input .xlsx file path")
    parser.add_argument("output", help="Output JSON file path")
    parser.add_argument(
        "--column",
        default="sentences",
        help="Header name of the Excel column containing text sentences",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name. If omitted, uses the first sheet.",
    )
    args = parser.parse_args()

    count = excel_to_posts_json(Path(args.input), Path(args.output), args.column, args.sheet)
    print(f"Created {args.output} with {count} posts.")


if __name__ == "__main__":
    main()
